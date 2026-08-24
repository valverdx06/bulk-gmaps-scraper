import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR
from utils.logger import logger

def export_data(results: list[dict], filename_prefix: str) -> tuple[str, str]:
    """
    Exports the scraped business lead results to Excel (.xlsx) and CSV (.csv) files.
    Applies professional English column headers, OpenPyXL styling, and UTF-8 BOM encoding.
    """
    if not results:
        logger.warning("No lead results found to export.")
        return "", ""

    # Create DataFrame
    df = pd.DataFrame(results)

    # 19 Enriched English Column Mapping
    column_mapping = {
        "name": "Business Name",
        "category": "Category",
        "rating": "Rating",
        "reviews_count": "Reviews Count",
        "phone": "Phone Number",
        "email": "Email Address",
        "website": "Website URL",
        "address": "Street Address",
        "city": "City",
        "postal_code": "Postal Code",
        "country": "Country",
        "latitude": "Latitude",
        "longitude": "Longitude",
        "facebook": "Facebook Profile",
        "instagram": "Instagram Profile",
        "linkedin": "LinkedIn Profile",
        "twitter": "Twitter Profile",
        "place_id": "Place ID",
        "url": "Google Maps Link"
    }

    # Filter and reorder columns according to mapping
    existing_cols = [col for col in column_mapping.keys() if col in df.columns]
    df = df[existing_cols]
    df.rename(columns=column_mapping, inplace=True)

    # Clean filename prefix
    clean_prefix = "".join([c if c.isalnum() or c in (" ", "_", "-") else "_" for c in filename_prefix]).strip()
    clean_prefix = clean_prefix.replace(" ", "_")

    excel_path = os.path.join(OUTPUT_DIR, f"{clean_prefix}.xlsx")
    csv_path = os.path.join(OUTPUT_DIR, f"{clean_prefix}.csv")

    # 1. Export to CSV (UTF-8 BOM for seamless Excel opening)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    logger.info(f"CSV report successfully saved to: {csv_path}")

    # 2. Export to Excel (.xlsx) with OpenPyXL styling
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        worksheet = writer.sheets["Leads"]

        # Header styling (Navy Blue fill, white bold font)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Auto-adjust column widths based on content
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Min width 15, Max width 55
            adjusted_width = min(max(max_len + 4, 15), 55)
            worksheet.column_dimensions[col_letter].width = adjusted_width

    logger.info(f"Excel report successfully saved to: {excel_path}")
    return excel_path, csv_path
